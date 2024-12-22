import requests
from bs4 import BeautifulSoup
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
import time
from plyer import notification
import gc
import threading
import tkinter as tk
from tkinter import ttk
import webbrowser
import re

# ----- Selenium関連 -----
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options

# ===== 複数カテゴリURL設定 =====
CHECK_URLS = [
    "https://kabutan.jp/news/marketnews/?category=3",
    "https://kabutan.jp/news/marketnews/?category=10",
    "https://kabutan.jp/news/marketnews/?category=2",
    "https://kabutan.jp/news/marketnews/?category=8",
    "https://kabutan.jp/news/marketnews/?category=9"
]

# URLごとの「前回取得した最新タイトル」記録ファイル
url_to_lastfile = {}
for url in CHECK_URLS:
    cat = url.split("category=")[-1]
    url_to_lastfile[url] = f"last_article_{cat}.txt"

EXCEL_FILE = "articles.xlsx"
RESULT_FONT = ("TkDefaultFont", 13)
BODY_FONT = ("TkDefaultFont", 13)

stop_flag = False
interval_minutes = 1
worker_thread = None

# 正規表現
CODE_PATTERN = re.compile(r"<(\d{4})>")
COLOR_PATTERN = re.compile("(赤字|黒字)")

def get_all_news(url):
    """
    指定URLからすべてのニュース情報を取得
    (ニュース時刻, カテゴリ, タイトル, フルURL) のリストを返す
    """
    try:
        response = requests.get(url)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        
        news_list = []
        trs = soup.select("table.s_news_list tr")
        for tr in trs:
            time_el = tr.select_one("td.news_time time")
            news_time = time_el.get_text(strip=True) if time_el else ""
            
            category_div = tr.select_one("td > div.newslist_ctg")
            category = category_div.get_text(strip=True) if category_div else ""
            
            link = tr.select_one("td > a")
            title = link.get_text(strip=True) if link else ""
            href = link.get("href", "") if link else ""
            full_url = "https://kabutan.jp" + href if href.startswith("/") else href
            
            if title and full_url:
                news_list.append((news_time, category, title, full_url))
        
        del soup, response
        return news_list
    except Exception as e:
        print(f"エラーが発生しました (URL: {url}): {e}")
        return []

def read_last_record(file_path):
    """
    前回取得した最新記事タイトルをテキストファイルから読み込む
    """
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read().strip()
    return None

def write_last_record(file_path, record):
    """
    最新の記事タイトルをテキストファイルへ書き込み
    """
    with open(file_path, 'w', encoding='utf-8') as f:
        f.write(record)

def append_to_excel(file_path, news_list):
    """
    ニュース情報のリストをExcelに追記
    (ニュースごとに DateTimeChecked, NewsTime, Category, Title, URL)
    """
    if not news_list:
        return

    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["DateTimeChecked", "NewsTime", "Category", "Title", "URL"])
    
    for news in news_list:
        news_time, category, title, url = news
        ws.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            news_time,
            category,
            title,
            url
        ])
    
    wb.save(file_path)
    del wb, ws

def show_notification(title, message):
    """
    デスクトップ通知を表示
    """
    notification.notify(
        title=title,
        message=message,
        timeout=5
    )

def open_sbi_with_code(code_list):
    """
    SBI証券サイトを開き、account.txtに記載のID/PWでログインし、
    株価検索フォームに銘柄コードを入力して検索する。
    ブラウザを閉じず、Seleniumの制御も解除
    """
    if not code_list:
        print("銘柄コードがありません。操作を中断します。")
        return

    code = code_list[0]
    sbi_url = "https://site2.sbisec.co.jp/ETGate/"

    if not os.path.exists("account.txt"):
        print("account.txt が存在しません。ID/PWを設定してください。")
        return
    with open("account.txt", "r", encoding="utf-8") as f:
        lines = f.read().splitlines()
    if len(lines) < 2:
        print("account.txt にIDとPWを正しく設定してください。(1行目:ID, 2行目:PW)")
        return
    user_id = lines[0].strip()
    user_pw = lines[1].strip()

    chrome_options = Options()
    chrome_options.add_experimental_option("detach", True)
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--disable-software-rasterizer")
    chrome_options.add_argument("--start-maximized")

    driver = webdriver.Chrome(options=chrome_options)
    try:
        driver.get(sbi_url)
        time.sleep(2)

        user_box = driver.find_element(By.NAME, "user_id")
        pass_box = driver.find_element(By.NAME, "user_password")
        user_box.send_keys(user_id)
        pass_box.send_keys(user_pw)
        login_btn = driver.find_element(By.NAME, "ACT_login")
        login_btn.click()

        time.sleep(3)
        try:
            stock_search_box = driver.find_element(By.NAME, "i_stock_sec")
            stock_search_box.send_keys(code)
            driver.find_element(By.CSS_SELECTOR, "#srchK > a").click()
        except Exception as e:
            print("株価検索フォーム操作失敗:", e)
            return

        print(f"銘柄コード {code} を検索しました。")
    except Exception as e:
        print("SBI証券での操作中にエラー:", e)

def show_body_window(url):
    """
    記事本文を別ウィンドウで表示
    - brタグ → 改行
    - 「。」 → 改行
    - 取得銘柄コード表示 & 証券会社ページへ遷移ボタン
    """
    try:
        resp = requests.get(url)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, 'html.parser')
        
        body_elem = soup.select_one(".body")
        if body_elem:
            #for br in body_elem.find_all("br"):
            #    br.replace_with("\n")
            body_text = body_elem.get_text()
            body_text = body_text.replace("。", "。\n")
        else:
            body_text = "本文を取得できませんでした。"
        del soup, resp
    except Exception as e:
        body_text = f"エラーが発生しました: {e}"

    win = tk.Toplevel(root)
    win.title("記事本文")

    content_frame = ttk.Frame(win, padding=8)
    content_frame.grid(sticky=(tk.W, tk.E, tk.N, tk.S))

    canvas = tk.Canvas(content_frame, width=1000, height=600)
    canvas.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)

    scrollbar = ttk.Scrollbar(content_frame, orient=tk.VERTICAL, command=canvas.yview)
    scrollbar.grid(row=0, column=3, sticky=(tk.N, tk.S))
    canvas.configure(yscrollcommand=scrollbar.set)

    body_frame = ttk.Frame(canvas)
    canvas.create_window((0,0), window=body_frame, anchor='nw')

    lines = body_text.split("\n")
    for line in lines:
        line_label = tk.Label(body_frame, text=line, font=BODY_FONT, justify="left")
        line_label.pack(anchor='w', pady=2, padx=5)

    def on_configure(event):
        canvas.config(scrollregion=canvas.bbox("all"))
    body_frame.bind("<Configure>", on_configure)

    # 銘柄コード
    codes = CODE_PATTERN.findall(body_text)
    if codes:
        codes_label = tk.Label(content_frame, text=f"取得した銘柄コード: {', '.join(codes)}", fg="green", font=BODY_FONT)
        codes_label.grid(row=1, column=0, columnspan=3, pady=5, sticky=tk.W)
    else:
        codes_label = tk.Label(content_frame, text="銘柄コードは検出されませんでした", fg="gray", font=BODY_FONT)
        codes_label.grid(row=1, column=0, columnspan=3, pady=5, sticky=tk.W)

    def open_original():
        webbrowser.open(url)
    open_button = tk.Button(content_frame, text="元ページを開く", command=open_original, bg="blue", fg="white", font=BODY_FONT)
    open_button.grid(row=2, column=0, pady=10, sticky=tk.W)

    def open_sbi():
        open_sbi_with_code(codes)
    sbi_button = tk.Button(content_frame, text="証券会社のページを開く", command=open_sbi, bg="green", fg="white", font=BODY_FONT)
    sbi_button.grid(row=2, column=1, pady=10, padx=10, sticky=tk.W)

    win.columnconfigure(0, weight=1)
    win.rowconfigure(0, weight=1)
    content_frame.columnconfigure(0, weight=1)
    content_frame.rowconfigure(0, weight=1)

def format_result_text(parent, text):
    """
    1行テキストから「赤字」を赤、「黒字」を青で表示
    """
    pos = 0
    matches = list(COLOR_PATTERN.finditer(text))
    if not matches:
        lbl = tk.Label(parent, text=text, font=RESULT_FONT)
        lbl.pack(side='left', padx=0)
        return

    for m in matches:
        keyword = m.group(0)
        color = "red" if keyword == "赤字" else "blue"
        start = m.start()
        if start > pos:
            segment = text[pos:start]
            lbl = tk.Label(parent, text=segment, font=RESULT_FONT)
            lbl.pack(side='left', padx=0)
        seg_lbl = tk.Label(parent, text=keyword, font=RESULT_FONT, fg=color)
        seg_lbl.pack(side='left', padx=0)
        pos = m.end()
    if pos < len(text):
        remainder = text[pos:]
        lbl = tk.Label(parent, text=remainder, font=RESULT_FONT)
        lbl.pack(side='left', padx=0)

# 行フレームをリストで管理 → 新着を上に表示
lines_list = []

def print_result(msg, url=None):
    """
    結果表示用エリアに新しい行を「上」に表示する。
    """
    def insert_line():
        line_frame = ttk.Frame(results_frame)
        # 新規行をリスト先頭に追加
        lines_list.insert(0, line_frame)

        # 中身を構築
        format_result_text(line_frame, msg)
        if url:
            def on_click():
                show_body_window(url)
            btn = tk.Button(line_frame, text="記事を見る", command=on_click, fg="blue", font=RESULT_FONT)
            btn.pack(side='left', padx=5)

        # 既存フレームをいったん pack_forget
        for lf in results_frame.winfo_children():
            lf.pack_forget()

        # lines_list の先頭から順に pack
        for lf in lines_list:
            lf.pack(fill='x', pady=2, anchor='w')

        results_frame.update_idletasks()
        canvas.config(scrollregion=canvas.bbox("all"))

    results_frame.after(0, insert_line)

def set_status(status):
    """
    ステータスラベルを更新する。スレッドセーフな方法で実行。
    """
    def update_label():
        status_label.config(text=status)
    root.after(0, update_label)

def check_for_update_on_url(url):
    """
    URLごとに最新ニュースをチェック
    新しい記事があればすべて取得し、Excelに追加
    last_articleがない場合は最新1記事のみ取得
    """
    news_list = get_all_news(url)
    if not news_list:
        # 記事が取得できなかった場合のみ表示
        print_result(f"{datetime.now()}: {url} 記事を取得できませんでした。")
        return

    last_record = read_last_record(url_to_lastfile[url])
    new_articles = []

    if last_record:
        # `last_record` 以降に追加された記事を取得
        for news in news_list:
            if news[2] == last_record:
                break
            new_articles.append(news)
    else:
        # `last_record` が存在しない場合、最新の1記事のみ取得
        if news_list:
            new_articles = [news_list[0]]

    if new_articles:
        # 新しい記事を逆順でExcelに追加（古い記事から順に追加）
        new_articles.reverse()
        append_to_excel(EXCEL_FILE, new_articles)
        if last_record:
            show_notification("新着記事", f"{len(new_articles)} 件の新しい記事が更新されました。")
        else:
            show_notification("新着記事", "最新の記事が取得されました。")
        # 最新の記事を `last_record` として保存
        latest_article = new_articles[-1][2]
        write_last_record(url_to_lastfile[url], latest_article)
        # 新しい記事を結果表示エリアに追加
        for news in new_articles:
            news_time, category, title, full_url = news
            msg = f"{news_time} / {category} / {title}"
            print_result(f"{news_time}: [{category}] {title}", url=full_url)
    else:
        # 更新がない場合は何も表示しない
        pass  # 仕様により表示しない

    del news_list, new_articles, last_record
    gc.collect()

def check_for_update_all_urls():
    """
    複数URL(CHECK_URLS)をまとめてチェック
    """
    for url in CHECK_URLS:
        check_for_update_on_url(url)

    now = datetime.now()
    now_ymd = "{0:%Y/%m/%d %H:%M}".format(now)
    print_result(f"{now_ymd}----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------")

def scraping_worker():
    global stop_flag, interval_minutes
    while not stop_flag:
        check_for_update_all_urls()
        for _ in range(interval_minutes * 60):
            if stop_flag:
                break
            time.sleep(1)
    set_status("停止中")

def start_scraping():
    """
    スクレイピング開始ボタン押下時 or 初期処理で呼び出し
    """
    global stop_flag, worker_thread, interval_minutes
    try:
        val = interval_entry.get().strip()
        if val:
            interval_minutes = int(val)
        else:
            interval_minutes = 1
    except ValueError:
        interval_minutes = 1

    stop_flag = False
    if worker_thread is None or not worker_thread.is_alive():
        worker_thread = threading.Thread(target=scraping_worker, daemon=True)
        worker_thread.start()
        set_status(f"実行中 (間隔: {interval_minutes}分)")
    else:
        # 既に実行中の場合、何もしない
        pass

def stop_scraping():
    """
    スクレイピング停止ボタン押下時
    """
    global stop_flag
    stop_flag = True
    set_status("停止中")

def init_excel_display():
    """
    アプリ起動時:
    1) Excelが存在・データがあれば結果表示へ日付順降順で表示
    2) なければ「データがありません」等を表示
    3) いずれでもその後スクレイピング開始
    """
    if not os.path.exists(EXCEL_FILE):
        print_result("Excelファイルがありません。記事データがないためスクレイピングを開始します。")
        start_scraping()
        return

    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))  # 2行目以降データ

        if len(rows) == 0:
            print_result("Excelに記事データがありません。スクレイピングを開始します。")
            wb.close()
            start_scraping()
            return

        # 各行を DateTimeChecked に基づいて降順にソート
        # DateTimeChecked は "YYYY-MM-DD HH:MM:SS" の形式を想定
        def parse_datetime(row):
            dt_str = row[0]
            try:
                return datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
            except:
                return datetime.min  # パース失敗時は最小値を返す

        sorted_rows = sorted(rows, key=parse_datetime, reverse=True)

        # ソートされた順に表示
        for row in sorted_rows:
            dt_checked, news_time, cat, title, url = row

            msg = f"{news_time}: [{cat}]  {title}"
            print_result(msg, url=url)

        wb.close()
    except Exception as e:
        print_result(f"Excelファイルの読み込み中にエラーが発生しました: {e}")
    
    start_scraping()

# ==================== GUI構築 ====================
root = tk.Tk()
root.title("ニューススクレイパー (Excel先読み & 日付順降順表示)")

frame = ttk.Frame(root, padding=20)
frame.grid(sticky=(tk.W, tk.E, tk.N, tk.S))

# --- ステータスラベルの追加 ---
status_label = ttk.Label(frame, text="停止中", font=("TkDefaultFont", 12), foreground="blue")
status_label.grid(row=0, column=0, columnspan=4, padx=5, pady=5, sticky=tk.W)

# --- コントロール部分 ---
interval_label = ttk.Label(frame, text="間隔(分):")
interval_label.grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)

interval_entry = ttk.Entry(frame, width=10)
interval_entry.grid(row=1, column=1, padx=5, pady=5, sticky=(tk.W, tk.E))
interval_entry.insert(0, "1")

start_button = tk.Button(frame, text="実行", bg="blue", fg="white",
                         command=start_scraping, width=10, font=RESULT_FONT)
start_button.grid(row=1, column=2, padx=5, pady=5, sticky=(tk.W))

stop_button = tk.Button(frame, text="停止", bg="red", fg="white",
                        command=stop_scraping, width=10, font=RESULT_FONT)
stop_button.grid(row=1, column=3, padx=5, pady=5, sticky=(tk.W))

# --- 結果表示エリア ---
canvas = tk.Canvas(frame, width=800, height=400)
canvas.grid(row=2, column=0, columnspan=4, padx=5, pady=5, sticky=(tk.W, tk.E, tk.N, tk.S))

scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=canvas.yview)
scrollbar.grid(row=2, column=4, sticky=(tk.N, tk.S))
canvas.configure(yscrollcommand=scrollbar.set)

results_frame = ttk.Frame(canvas)
canvas.create_window((0,0), window=results_frame, anchor='nw')

def on_configure(event):
    canvas.config(scrollregion=canvas.bbox("all"))

results_frame.bind("<Configure>", on_configure)

# 管理用リストの設定
lines_list = []

root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)
frame.columnconfigure(1, weight=1)
frame.rowconfigure(2, weight=1)

# アプリ起動時にExcel読み込み＆初期表示 → スクレイピング開始
root.after(0, init_excel_display)

root.mainloop()
